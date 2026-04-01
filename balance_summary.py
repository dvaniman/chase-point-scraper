"""
Build a balance summary from balance_history: current balance plus 30/60/90-day lookbacks.
Reads output/balance_history.xlsx (or .csv), writes output/balance_summary.xlsx.

Columns: Account Name, Current Balance, Balance 30d, Balance 60d, Balance 90d.
For 30/60/90d we use the snapshot closest to that many days ago, only if within +/-5 days;
otherwise the cell is left blank.

Run: python balance_summary.py
"""

import csv
from pathlib import Path
from datetime import datetime, timedelta
from typing import List

OUTPUT_DIR = Path(__file__).resolve().parent / "output"
HISTORY_FILENAME = "balance_history.xlsx"
SUMMARY_FILENAME = "balance_summary.xlsx"


def load_config():
    try:
        import yaml
        config_path = Path(__file__).resolve().parent / "config.yaml"
        if config_path.exists():
            with open(config_path, "r", encoding="utf-8") as f:
                return yaml.safe_load(f) or {}
    except Exception:
        pass
    return {}


def load_balance_history(path: Path) -> List[dict]:
    """Load rows from balance_history.xlsx or .csv. Each row: Snapshot At, Account Name, Points Balance."""
    if not path.exists():
        return []
    rows = []
    if path.suffix.lower() == ".csv":
        with open(path, "r", encoding="utf-8", newline="") as f:
            r = csv.DictReader(f)
            for row in r:
                rows.append(dict(row))
        return rows
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb["Balance history"] if "Balance history" in wb.sheetnames else wb.active
        headers = []
        for c in range(1, ws.max_column + 1):
            v = ws.cell(1, c).value
            headers.append(str(v).strip() if v else "")
        for row in range(2, ws.max_row + 1):
            r = {}
            for col, h in enumerate(headers, 1):
                if h:
                    r[h] = ws.cell(row, col).value
            if r:
                rows.append(r)
        wb.close()
    except Exception:
        pass
    return rows


def parse_snapshot_at(s: str):
    """Parse Snapshot At to datetime (date used for day math)."""
    if s is None or not str(s).strip():
        return None
    s = str(s).strip()
    for fmt, size in (("%Y-%m-%dT%H:%M:%S", 19), ("%Y-%m-%d %H:%M:%S", 19), ("%Y-%m-%d", 10)):
        if len(s) >= size:
            try:
                return datetime.strptime(s[:size], fmt)
            except ValueError:
                continue
    return None


def balance_at_target_days_ago(records: List[tuple], target_days: int, tolerance_days: int = 5):
    """
    From list of (date, balance), return balance for the snapshot whose date is closest to
    target_days ago, only if within +/- tolerance_days; else None.
    """
    if not records:
        return None
    today = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
    target = today - timedelta(days=target_days)
    low = today - timedelta(days=target_days + tolerance_days)
    high = today - timedelta(days=target_days - tolerance_days)
    best_balance = None
    best_diff = None
    for dt, bal in records:
        if bal is None:
            continue
        try:
            b = int(float(bal))
        except (TypeError, ValueError):
            continue
        d = (dt.replace(hour=0, minute=0, second=0, microsecond=0) if isinstance(dt, datetime) else dt)
        if not (low <= d <= high):
            continue
        diff = abs((d - target).days)
        if best_diff is None or diff < best_diff:
            best_diff = diff
            best_balance = b
    return best_balance


def build_summary(history_path: Path) -> List[dict]:
    """One row per account: Account Name, Current Balance, Balance 30d, Balance 60d, Balance 90d."""
    raw = load_balance_history(history_path)
    if not raw:
        return []
    # Normalize column names
    snap_key = None
    account_key = None
    balance_key = None
    for k in raw[0].keys():
        if "snapshot" in k.lower() or "at" in k.lower():
            snap_key = k
        if "account" in k.lower() and "name" in k.lower():
            account_key = k
        if "balance" in k.lower() and "point" in k.lower():
            balance_key = k
    if not snap_key:
        snap_key = "Snapshot At"
    if not account_key:
        account_key = "Account Name"
    if not balance_key:
        balance_key = "Points Balance"

    by_account = {}
    for r in raw:
        acc = (r.get(account_key) or "").strip()
        if not acc:
            continue
        dt = parse_snapshot_at(r.get(snap_key))
        if dt is None:
            continue
        try:
            bal = int(float(r.get(balance_key) or 0))
        except (TypeError, ValueError):
            continue
        if acc not in by_account:
            by_account[acc] = []
        by_account[acc].append((dt, bal))

    summary = []
    for acc in sorted(by_account.keys()):
        recs = by_account[acc]
        recs.sort(key=lambda x: x[0], reverse=True)
        current = recs[0][1] if recs else None
        bal_30 = balance_at_target_days_ago(recs, 30, 5)
        bal_60 = balance_at_target_days_ago(recs, 60, 5)
        bal_90 = balance_at_target_days_ago(recs, 90, 5)
        summary.append({
            "Account Name": acc,
            "Current Balance": current,
            "Balance 30d": bal_30,
            "Balance 60d": bal_60,
            "Balance 90d": bal_90,
        })
    return summary


def _totals_row(rows: List[dict]) -> dict:
    totals = {"Account Name": "Total"}
    for key in ("Current Balance", "Balance 30d", "Balance 60d", "Balance 90d"):
        totals[key] = sum((r.get(key) or 0) for r in rows if isinstance(r.get(key), (int, float)))
    return totals


def write_summary_csv(rows: List[dict], path: Path) -> None:
    if not rows:
        return
    headers = ["Account Name", "Current Balance", "Balance 30d", "Balance 60d", "Balance 90d"]
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=headers, extrasaction="ignore")
        w.writeheader()
        for r in rows:
            w.writerow({k: ("" if v is None else v) for k, v in r.items()})
        w.writerow({k: ("" if v is None else v) for k, v in _totals_row(rows).items()})


def write_summary_xlsx(rows: List[dict], path: Path) -> None:
    if not rows:
        return
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        write_summary_csv(rows, path.with_suffix(".csv"))
        return
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Summary"
    headers = ["Account Name", "Current Balance", "Balance 30d", "Balance 60d", "Balance 90d"]
    header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = header_fill
        c.font = header_font
    for row_idx, r in enumerate(rows, 2):
        ws.cell(row=row_idx, column=1, value=r.get("Account Name", ""))
        for col, key in enumerate(["Current Balance", "Balance 30d", "Balance 60d", "Balance 90d"], 2):
            v = r.get(key)
            ws.cell(row=row_idx, column=col, value=v if v is not None else "")
    totals = _totals_row(rows)
    total_row = len(rows) + 2
    ws.cell(row=total_row, column=1, value="Total")
    for col, key in enumerate(["Current Balance", "Balance 30d", "Balance 60d", "Balance 90d"], 2):
        v = totals.get(key)
        cell = ws.cell(row=total_row, column=col, value=v if v is not None else "")
        cell.font = Font(bold=True)
    wb.save(path)


def main():
    config = load_config()
    out_cfg = config.get("output", {})
    history_name = (out_cfg.get("balance_history_file") or "balance_history").strip()
    if not history_name.endswith(".xlsx") and not history_name.endswith(".csv"):
        history_name += ".xlsx"
    history_path = OUTPUT_DIR / history_name
    if not history_path.exists():
        history_path = OUTPUT_DIR / "balance_history.csv"
    if not history_path.exists():
        print(f"Balance history not found at {history_path}. Run the scraper first to build balance_history.")
        return
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    summary_path = OUTPUT_DIR / SUMMARY_FILENAME
    rows = build_summary(history_path)
    if not rows:
        print("No balance history rows found. Run the scraper to record balances.")
        return
    write_summary_xlsx(rows, summary_path)
    print(f"Wrote {len(rows)} account(s) to {summary_path}")


if __name__ == "__main__":
    main()
