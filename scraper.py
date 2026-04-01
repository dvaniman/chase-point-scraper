"""
Chase points activity scraper.
Log in with credentials from .env; scrapes points data and exports to CSV/Excel
with columns: Date, Account Name, Payee, Type, EarnX, Dollars, Points.
"""

import os
import re
import sys
import csv
from collections import Counter
from pathlib import Path
from datetime import datetime

import yaml
from dotenv import load_dotenv
from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeout

# Load env first
load_dotenv()

# Default config path
CONFIG_PATH = Path(__file__).resolve().parent / "config.yaml"
OUTPUT_DIR = Path(__file__).resolve().parent / "output"
SESSION_DIR = Path(__file__).resolve().parent / "session"


def _session_paths_for_account(account_name: str, config: dict = None) -> tuple:
    """Return (profile_dir, session_file). Profile is in LOCALAPPDATA so 2FA cookies persist.
    Accounts in the same profile_group share one profile (2FA once per group)."""
    group = (config or {}).get("profile_group", {}).get(account_name) or account_name
    safe = re.sub(r'[<>:"/\\|?*]', "_", group).strip() or "account"
    local_base = Path(os.environ.get("LOCALAPPDATA", os.path.expanduser("~")))
    profile_dir = local_base / "ChaseScraper" / "profiles" / safe
    session_file = local_base / "ChaseScraper" / "sessions" / f"chase_state_{safe}.json"
    return profile_dir, session_file


def load_config():
    with open(CONFIG_PATH, "r", encoding="utf-8") as f:
        return yaml.safe_load(f)


def _prompt_browser(msg: str, then: str = "Then press ENTER in this terminal to continue.") -> None:
    """Print a clear prompt so the user knows when to act in the browser."""
    sep = "=" * 60
    print()
    print(sep)
    print("  >>> DO IN BROWSER:", msg)
    print("  >>>", then)
    print(sep)
    print()
    sys.stdout.flush()


def prompt_account_menu(config):
    """Show a numbered menu of account choices. Returns a list of account names (one or all)."""
    choices = config.get("account_choices") or [
        "Ink Pref", "Chase SR", "Freedom-kev", "Ink Cash", "Ink Unlt", "Freedom-dsv", "Sapphire Pref",
    ]
    print()
    print("Select account (credentials and session are per-account):")
    print("-" * 40)
    for i, name in enumerate(choices, 1):
        print(f"  {i}. {name}")
    print(f"  {len(choices) + 1}. All (1-{len(choices)} in succession, sign out after each)")
    print("-" * 40)
    sys.stdout.flush()
    while True:
        try:
            raw = input("Enter number (1-{}): ".format(len(choices) + 1)).strip()
            n = int(raw)
            if 1 <= n <= len(choices):
                return [choices[n - 1]]
            if n == len(choices) + 1:
                return choices
        except (ValueError, EOFError):
            pass
        print("Invalid. Enter a number from 1 to {}.".format(len(choices) + 1))
        sys.stdout.flush()


def get_credentials(account_name: str, config: dict):
    """Get Chase username and password for the selected account from env (see config account_credentials)."""
    creds = (config.get("account_credentials") or {}).get(account_name)
    if creds:
        username = os.environ.get(creds.get("username_env") or "")
        password = os.environ.get(creds.get("password_env") or "")
    else:
        username = os.environ.get("CHASE_USERNAME")
        password = os.environ.get("CHASE_PASSWORD")
    if not username or not password:
        raise SystemExit(
            f"Set credentials for account '{account_name}' in .env. "
            "See config.yaml account_credentials for the env var names (e.g. CHASE_USER_INK_PREF, CHASE_PASSWORD_INK_PREF)."
        )
    return username, password


def normalize_header(text):
    """Map common header labels to our canonical names."""
    t = (text or "").strip().lower()
    if "date" in t:
        return "Date"
    if "payee" in t or "merchant" in t or "description" in t:
        return "Payee"
    if "type" in t or "category" in t:
        return "Type"
    if "earn" in t or "multiplier" in t or "earnx" in t:
        return "EarnX"
    if "dollar" in t or "amount" in t:
        return "Dollars"
    if "point" in t:
        return "Points"
    return None


def _parse_rows_with_header_map(header_cell_texts, data_rows, get_cell_texts):
    """Build header index -> canonical name, then extract row dicts. get_cell_texts(row) returns list of str."""
    header_map = []
    for i, h in enumerate(header_cell_texts):
        canonical = normalize_header(h)
        if canonical:
            header_map.append((i, canonical))
    if not header_map:
        return []
    out = []
    for row_el in data_rows:
        cell_texts = get_cell_texts(row_el)
        if not cell_texts:
            continue
        row_data = {canonical: (cell_texts[idx] if idx < len(cell_texts) else "") for idx, canonical in header_map}
        if any(row_data.values()):
            out.append(row_data)
    return out


def _parse_balance_int(text):
    """Extract a points balance integer from text like '125,430 pts' or '125430'."""
    if not text:
        return None
    m = re.search(r"([\d,]+)\s*(?:points?|pts)\b", text, re.I)
    if m:
        try:
            return int(m.group(1).replace(",", ""))
        except ValueError:
            pass
    m = re.search(r"\b([\d]{1,3}(?:,\d{3})+)\b", text)
    if m:
        try:
            return int(m.group(1).replace(",", ""))
        except ValueError:
            pass
    return None


def _balance_js_selector():
    """
    JS: get balance from selector page (shadow DOM). Returns int or { balance: null, itemCount, labels, snippet } for debug.
    Tries multiple patterns: "Available Points: N pts", "available points", "N pts" near "available", etc.
    """
    return r"""
    (last4) => {
        function allText(node) {
            if (!node) return '';
            if (node.nodeType === 3) return node.textContent || '';
            let s = '';
            if (node.shadowRoot) s += allText(node.shadowRoot);
            const ch = node.childNodes;
            if (ch) for (let i = 0; i < ch.length; i++) s += allText(ch[i]);
            return s;
        }
        const L = (last4 || '').trim();
        const items = document.querySelectorAll('mds-list-item');
        const labels = [];
        for (let i = 0; i < items.length; i++) {
            const lab = (items[i].getAttribute('label') || '');
            labels.push(lab ? lab.substring(0, 80) : '(no label)');
        }
        let snippet = '';
        for (let i = 0; i < items.length; i++) {
            const el = items[i];
            const lab = (el.getAttribute('label') || '');
            if (L && lab.indexOf(L) === -1) continue;
            const txt = allText(el).replace(/\s+/g, ' ');
            snippet = txt.substring(0, 1500);
            const patterns = [
                /Available\s+Points:\s*([\d,]+)\s*pts?/i,
                /available\s+points:\s*([\d,]+)/i,
                /([\d,]+)\s*pts?\s*available/i,
                /([\d,]{2,})\s*pts?/,
                /points?\s*[:\s]*([\d,]+)/i
            ];
            for (const re of patterns) {
                const m = txt.match(re);
                if (m) {
                    const num = parseInt(m[1].replace(/,/g, ''), 10);
                    if (num >= 1 && num < 1e8) return num;
                }
            }
        }
        const body = allText(document.body).replace(/\s+/g, ' ');
        if (!snippet) snippet = body.substring(0, 2000);
        for (const re of [/Available\s+Points:\s*([\d,]+)\s*pts?/gi, /([\d,]+)\s*pts?\s*available/gi]) {
            let m;
            while ((m = re.exec(body)) !== null) {
                const num = parseInt(m[1].replace(/,/g, ''), 10);
                if (num >= 100 && num < 1e8) return num;
            }
        }
        return { balance: null, itemCount: items.length, labels: labels, snippet: snippet };
    }
    """


def scrape_balance_from_selector_page(page, last4: str):
    """
    On the account selector page, find the card for last4 and parse Available Points.
    Chase puts the balance in the mds-list-item **description** attribute, e.g.
    description="Available Points: 87,651 pts" (not shadow-DOM text).
    """
    if not (last4 or "").strip():
        return None

    def parse_balance_from_text(text):
        if not text:
            return None
        m = re.search(r"Available\s+Points:\s*([\d,]+)\s*pts?", text, re.I)
        if m:
            try:
                return int(m.group(1).replace(",", ""))
            except ValueError:
                pass
        return _parse_balance_int(text)

    # Primary: description attribute on the matching card (Chase light DOM)
    try:
        card = page.locator(f'mds-list-item[label*="{last4}"]').first
        card.wait_for(state="attached", timeout=12000)
        desc = (card.get_attribute("description") or "").strip()
        n = parse_balance_from_text(desc)
        if n is not None:
            return n
    except Exception as e:
        print(f"[Balance debug] description attribute: {e}")
        sys.stdout.flush()

    try:
        page.wait_for_timeout(800)
        result = page.evaluate(_balance_js_selector(), last4)
        if result is None:
            pass
        elif isinstance(result, dict):
            if result.get("balance") is not None:
                try:
                    n = int(float(result["balance"]))
                    if n >= 1:
                        return n
                except (TypeError, ValueError):
                    pass
            snippet = result.get("snippet") or ""
            item_count = result.get("itemCount", 0)
            labels = result.get("labels") or []
            OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
            debug_path = OUTPUT_DIR / "balance_debug_selector.txt"
            debug_path.write_text(snippet, encoding="utf-8")
            print(
                f"[Balance debug] Selector page: {item_count} card(s), labels: {labels!r}. "
                f"Page text written to {debug_path}"
            )
            sys.stdout.flush()
        else:
            try:
                n = int(float(result))
                if n >= 1:
                    return n
            except (TypeError, ValueError):
                pass
    except Exception as e:
        print(f"[Balance debug] Selector page JS error: {e}")
        sys.stdout.flush()

    try:
        card = page.locator(f'mds-list-item[label*="{last4}"]').first
        card.wait_for(state="attached", timeout=5000)
        text = card.inner_text() or ""
        n = parse_balance_from_text(text)
        if n is not None:
            return n
    except Exception as e:
        print(f"[Balance debug] Selector page inner_text: {e}")
        sys.stdout.flush()

    try:
        content = page.content() or ""
        if last4 not in content:
            print(f"[Balance debug] last4 {last4!r} not found in page HTML.")
            sys.stdout.flush()
        else:
            for m in re.finditer(r"Available\s+Points:\s*([\d,]+)\s*pts?", content, re.I):
                try:
                    return int(m.group(1).replace(",", ""))
                except ValueError:
                    continue
    except Exception as e:
        print(f"[Balance debug] page.content(): {e}")
        sys.stdout.flush()
    return None


def scrape_balance_from_activity_page(page):
    """
    After activity loads, total points often appears once on page (including shadow DOM).
    """
    try:
        page.wait_for_timeout(500)
        result = page.evaluate(r"""() => {
            function allText(node) {
                if (!node) return '';
                if (node.nodeType === 3) return node.textContent || '';
                let s = '';
                if (node.shadowRoot) s += allText(node.shadowRoot);
                const ch = node.childNodes;
                if (ch) for (let i = 0; i < ch.length; i++) s += allText(ch[i]);
                return s;
            }
            const body = allText(document.body).replace(/\s+/g, ' ');
            const matches = [];
            let re = /Available\s+Points:\s*([\d,]+)\s*pts?/gi;
            let m;
            while ((m = re.exec(body)) !== null) {
                matches.push(parseInt(m[1].replace(/,/g, ''), 10));
            }
            if (matches.length >= 1) return Math.max.apply(null, matches);
            const reBal = /(balance|available|total|you have)[^0-9]{0,40}([\d,]{2,})\s*pts?/gi;
            const bal = [];
            while ((m = reBal.exec(body)) !== null) {
                const v = parseInt(m[2].replace(/,/g, ''), 10);
                if (v >= 100) bal.push(v);
            }
            if (bal.length === 1) return bal[0];
            if (bal.length > 1) return Math.max.apply(null, bal);
            return { balance: null, snippet: body.substring(0, 2000) };
        }""")
        if result is None:
            return None
        if isinstance(result, dict):
            OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
            debug_path = OUTPUT_DIR / "balance_debug_activity.txt"
            debug_path.write_text(result.get("snippet") or "", encoding="utf-8")
            print(f"[Balance debug] Activity page: no balance pattern. Wrote {debug_path}")
            sys.stdout.flush()
            return None
        try:
            n = int(float(result))
            if n >= 100:
                return n
        except (TypeError, ValueError):
            pass
    except Exception as e:
        print(f"[Balance debug] Activity page error: {e}")
        sys.stdout.flush()
    return None


def scrape_points_balance(page, config):
    """
    Read current Ultimate Rewards points balance from the page.
    Tries config selectors first, then contextual regex (balance / available / you have).
    Returns int or None.
    """
    selectors = config.get("selectors", {})
    raw = (selectors.get("points_balance") or "").strip()
    for sel in [s.strip() for s in raw.split(",") if s.strip()]:
        try:
            for el in page.query_selector_all(sel):
                t = (el.inner_text() or "").strip()
                if not t or len(t) > 500:
                    continue
                n = _parse_balance_int(t)
                if n is not None and n >= 1:
                    return n
        except Exception:
            continue
    try:
        text = page.inner_text() or ""
        # Prefer numbers near balance-related wording
        for m in re.finditer(
            r".{0,80}(\d{1,3}(?:,\d{3})+)\s*(?:points|pts)\b.{0,80}",
            text[:20000],
            re.I | re.DOTALL,
        ):
            chunk = m.group(0).lower()
            if any(
                w in chunk
                for w in (
                    "balance",
                    "available",
                    "total",
                    "you have",
                    "rewards balance",
                    "point balance",
                    "current",
                )
            ):
                try:
                    return int(m.group(1).replace(",", ""))
                except ValueError:
                    continue
        # Last resort: largest plausible account balance (exclude tiny per-txn amounts)
        candidates = []
        for m in re.finditer(r"(\d{1,3}(?:,\d{3})+)\s*(?:points|pts)\b", text[:12000], re.I):
            try:
                v = int(m.group(1).replace(",", ""))
                if v >= 1000:
                    candidates.append(v)
            except ValueError:
                continue
        if candidates:
            return max(candidates)
    except Exception:
        pass
    return None


def append_balance_snapshot(account_name: str, balance: int, config: dict) -> Path | None:
    """
    Append one row: when scraped, which account, points balance.
    Single file for all accounts — filter or pivot by Account Name in Excel.
    """
    out_cfg = config.get("output", {})
    if not out_cfg.get("record_balance_each_run", True):
        return None
    name = (out_cfg.get("balance_history_file") or "balance_history").strip()
    if not name.endswith(".xlsx") and not name.endswith(".csv"):
        name += ".xlsx"
    path = OUTPUT_DIR / name
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    now = datetime.now().isoformat(timespec="seconds")
    headers = ["Snapshot At", "Account Name", "Points Balance"]
    row = [now, account_name, balance]

    if path.suffix.lower() == ".csv":
        new_file = not path.exists()
        with open(path, "a", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            if new_file:
                w.writerow(headers)
            w.writerow(row)
        return path

    try:
        import openpyxl
    except ImportError:
        csv_path = path.with_suffix(".csv")
        new_file = not csv_path.exists()
        with open(csv_path, "a", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            if new_file:
                w.writerow(headers)
            w.writerow(row)
        return csv_path

    if path.exists():
        wb = openpyxl.load_workbook(path)
        if "Balance history" in wb.sheetnames:
            ws = wb["Balance history"]
        else:
            ws = wb.create_sheet("Balance history", 0)
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Balance history"
    if ws.max_row == 0:
        ws.append(headers)
    ws.append(row)
    wb.save(path)
    wb.close()
    return path


def ensure_balance_history_file(config: dict) -> Path:
    """
    Create balance_history file with headers if it doesn't exist, so the file
    is present in output/ even before the first balance is recorded.
    """
    out_cfg = config.get("output", {})
    if not out_cfg.get("record_balance_each_run", True):
        return OUTPUT_DIR / "balance_history.xlsx"
    name = (out_cfg.get("balance_history_file") or "balance_history").strip()
    if not name.endswith(".xlsx") and not name.endswith(".csv"):
        name += ".xlsx"
    path = OUTPUT_DIR / name
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    if path.exists():
        return path
    headers = ["Snapshot At", "Account Name", "Points Balance"]
    if path.suffix.lower() == ".csv":
        with open(path, "w", newline="", encoding="utf-8") as f:
            csv.writer(f).writerow(headers)
        return path
    try:
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Balance history"
        ws.append(headers)
        wb.save(path)
        wb.close()
    except ImportError:
        csv_path = path.with_suffix(".csv")
        with open(csv_path, "w", newline="", encoding="utf-8") as f:
            csv.writer(f).writerow(headers)
        return csv_path
    return path


def scrape_table_and_account(page, config):
    """Find activity table and optional account name on the current page."""
    selectors = config.get("selectors", {})
    account_name = "Chase"  # default

    # Try to get account/card name from page
    for sel in selectors.get("account_name", "").split(", "):
        sel = sel.strip()
        if not sel:
            continue
        try:
            el = page.query_selector(sel)
            if el:
                text = el.inner_text().strip()
                if text and len(text) < 200:
                    account_name = text
                    break
        except Exception:
            continue

    all_rows = []

    # Strategy 0: Chase Ultimate Rewards page (mds-list-item / transaction-details-item, not a table)
    try:
        items = page.query_selector_all("transaction-details-item mds-list-item, mds-list#posted-activity-list mds-list-item, [id='posted-activity-list'] mds-list-item")
        if not items:
            items = page.query_selector_all("mds-list-item[label][description]")
        for item in items:
            try:
                label = item.get_attribute("label") or ""
                desc = item.get_attribute("description") or ""
                sec_desc = item.get_attribute("secondary-description") or ""
                sec_label = item.get_attribute("secondary-label") or ""
            except Exception:
                continue
            if not label and not sec_label:
                continue
            # description is e.g. "Mar 1, 2026&lt;br/&gt;1.5% earn" -> date and earn type
            parts = desc.replace("&lt;", "<").replace("&gt;", ">").split("<br/>") if desc else ["", ""]
            date_str = (parts[0] or "").strip()
            earn_str = (parts[1] if len(parts) > 1 else "").strip()  # e.g. "1.5% earn" or "3% earn"
            # Normalize to EarnX: "1.5% earn" -> "1.5x", "3% earn" -> "3x"
            earn_x = ""
            if "%" in earn_str:
                m = re.search(r"([\d.]+)\s*%", earn_str)
                if m:
                    earn_x = m.group(1) + "x"
            if not earn_x and earn_str:
                earn_x = "1x"
            dollars = (sec_desc or "").replace("$", "").replace(",", "").strip()
            points = (sec_label or "").replace(" pts", "").replace("pts", "").replace(",", "").strip()
            row = {
                "Date": normalize_date_str(date_str) or date_str,
                "Payee": label,
                "Type": "Earn" if "earn" in earn_str.lower() else "",
                "EarnX": earn_x,
                "Dollars": dollars,
                "Points": points,
            }
            if row["Payee"] or row["Points"]:
                all_rows.append(row)
        if all_rows:
            return all_rows, account_name
    except Exception:
        pass

    # Strategy 1: HTML table with tr / th, td
    table_selectors = (selectors.get("activity_table") or "table").split(", ")
    for table_sel in table_selectors:
        table_sel = table_sel.strip()
        if not table_sel:
            continue
        try:
            tables = page.query_selector_all(table_sel)
        except Exception:
            tables = []
        for table in tables:
            try:
                rows = table.query_selector_all("tr")
            except Exception:
                rows = []
            if len(rows) < 2:
                continue
            header_cells = rows[0].query_selector_all("th, td")
            headers = [c.inner_text().strip() for c in header_cells]
            all_rows = _parse_rows_with_header_map(headers, rows[1:], lambda r: [c.inner_text().strip() for c in r.query_selector_all("td, th")])
            if all_rows:
                return all_rows, account_name

    # Strategy 2: ARIA grid ([role=grid] with [role=row] and [role=cell] / [role=gridcell])
    try:
        grids = page.query_selector_all("[role='grid']")
        for grid in grids:
            rows = grid.query_selector_all("[role='row']")
            if len(rows) < 2:
                continue
            header_cells = rows[0].query_selector_all("[role='columnheader'], [role='rowheader'], [role='cell'], [role='gridcell']")
            if not header_cells:
                header_cells = rows[0].query_selector_all("[role='cell'], [role='gridcell']")
            headers = [c.inner_text().strip() for c in header_cells]
            if not any(normalize_header(h) for h in headers):
                continue

            def get_cells(row_el):
                cells = row_el.query_selector_all("[role='cell'], [role='gridcell']")
                return [c.inner_text().strip() for c in cells]

            all_rows = _parse_rows_with_header_map(headers, rows[1:], get_cells)
            if all_rows:
                return all_rows, account_name
    except Exception:
        pass

    # Strategy 3: div/section with row-like children and header-like first row
    try:
        for tag in ["table", "div", "section"]:
            containers = page.query_selector_all(tag)
            for cont in containers:
                row_like = cont.query_selector_all("tr, [role='row'], [class*='row']")
                if len(row_like) < 2:
                    continue
                first = row_like[0]
                cells = first.query_selector_all("th, td, [role='cell'], [role='gridcell'], [class*='cell'], [class*='col']")
                if not cells:
                    continue
                headers = [c.inner_text().strip() for c in cells]
                if not any(normalize_header(h) for h in headers):
                    continue

                def get_cells_row(r):
                    cels = r.query_selector_all("td, th, [role='cell'], [role='gridcell'], [class*='cell'], [class*='col']")
                    return [x.inner_text().strip() for x in cels]

                all_rows = _parse_rows_with_header_map(headers, row_like[1:], get_cells_row)
                if all_rows:
                    return all_rows, account_name
    except Exception:
        pass

    return all_rows, account_name


def add_account_column(rows, account_name):
    """Ensure every row has 'Account Name'."""
    for r in rows:
        if "Account Name" not in r:
            r["Account Name"] = account_name
    return rows


def ensure_column_order(rows, column_order):
    """Return list of dicts with keys in column_order; missing keys get ''."""
    result = []
    for r in rows:
        result.append({k: r.get(k, "") for k in column_order})
    return result


def _account_file_path(account_name: str) -> Path:
    """Path to the ongoing xlsx file for this account (one file per account)."""
    safe = re.sub(r'[<>:"/\\|?*]', "_", account_name).strip() or "account"
    return OUTPUT_DIR / f"{safe}.xlsx"


# Canonical date format for storage and dedupe (YYYY-MM-DD)
DATE_FMT_CANONICAL = "%Y-%m-%d"
DATE_INPUT_FORMATS = [
    "%Y-%m-%d %H:%M:%S",  # 2025-09-15 00:00:00 (Excel datetime)
    "%Y-%m-%dT%H:%M:%S",
    "%Y-%m-%d",
    "%b %d, %Y",   # Mar 1, 2026
    "%B %d, %Y",   # March 1, 2026
    "%m/%d/%Y",    # 03/01/2026
    "%m-%d-%Y",    # 03-01-2026
    "%d/%m/%Y",    # 01/03/2026
    "%b %d, %y",   # Mar 1, 26
]


def normalize_date_str(s: str) -> str:
    """
    Parse common date strings and return YYYY-MM-DD for consistent sorting and dedupe.
    Handles Excel-style "2025-09-15 00:00:00". Returns original string if unparseable.
    """
    s = (s or "").strip()
    if not s:
        return s
    # Strip Excel serial or trailing time so we only parse the date part
    s_trim = s[:50].strip()
    if re.match(r"^\d{4}-\d{2}-\d{2}\s+\d{2}:\d{2}:\d{2}", s_trim):
        s_trim = s_trim[:10]
    for fmt in DATE_INPUT_FORMATS:
        try:
            dt = datetime.strptime(s_trim, fmt)
            return dt.strftime(DATE_FMT_CANONICAL)
        except ValueError:
            continue
    # Try stripping trailing year shorthand (e.g. "Mar 1, 20" -> assume 2020)
    m = re.match(r"^(\w+\s+\d{1,2}),\s*(\d{2})$", s.strip())
    if m:
        try:
            dt = datetime.strptime(f"{m.group(1)}, 20{m.group(2)}", "%b %d, %Y")
            return dt.strftime(DATE_FMT_CANONICAL)
        except ValueError:
            pass
    return s


def _date_sort_key(row: dict) -> tuple:
    """Sort key for date order (chronological = oldest first). Uses normalized date when possible."""
    s = str(row.get("Date", "")).strip()
    if not s:
        return (9999, 99, 99)
    normalized = normalize_date_str(s)
    if normalized != s and re.match(r"^\d{4}-\d{2}-\d{2}$", normalized):
        try:
            dt = datetime.strptime(normalized, DATE_FMT_CANONICAL)
            return (dt.year, dt.month, dt.day)
        except ValueError:
            pass
    for fmt in DATE_INPUT_FORMATS + [DATE_FMT_CANONICAL]:
        try:
            dt = datetime.strptime(s[:50], fmt)
            return (dt.year, dt.month, dt.day)
        except ValueError:
            continue
    return (9999, 99, 99)


def _row_key(row: dict) -> tuple:
    """Key for deduplication: same Date (normalized), Payee, Dollars, Points = same transaction."""
    date_val = str(row.get("Date", "")).strip()
    if date_val and re.match(r"^\d{4}-\d{2}-\d{2}$", normalize_date_str(date_val)):
        date_val = normalize_date_str(date_val)
    return (
        date_val,
        str(row.get("Payee", "")),
        str(row.get("Dollars", "")),
        str(row.get("Points", "")),
    )


def load_existing_rows(path: Path, column_order: list) -> list:
    """Load existing rows from an account xlsx (or csv fallback); return list of dicts with column_order keys."""
    load_path = path
    if not load_path.exists() and path.suffix.lower() == ".xlsx":
        csv_path = path.with_suffix(".csv")
        if csv_path.exists():
            load_path = csv_path
    if not load_path.exists():
        return []
    if load_path.suffix.lower() == ".xlsx":
        try:
            import openpyxl
            wb = openpyxl.load_workbook(load_path, read_only=True, data_only=True)
            ws = wb.active
            rows_iter = ws.iter_rows(min_row=1, values_only=True)
            header = next(rows_iter, None)
            if not header:
                wb.close()
                return []
            keys = [str(h) if h is not None else "" for h in header]
            out = []
            for row in rows_iter:
                r = dict(zip(keys, (v if v is not None else "" for v in row)))
                out.append({k: r.get(k, "") for k in column_order})
            wb.close()
            return out
        except Exception:
            return []
    # CSV fallback (e.g. old file)
    out = []
    with open(load_path, "r", encoding="utf-8", newline="") as f:
        reader = csv.DictReader(f)
        for r in reader:
            out.append({k: r.get(k, "") for k in column_order})
    return out


def backfill_balance_sheets(config=None):
    """
    One-time: copy latest balance from balance_history.xlsx into each account workbook
    that is missing a 'Balance' sheet (e.g. Ink Pref.xlsx).
    """
    config = config or load_config()
    out_cfg = config.get("output", {})
    name = (out_cfg.get("balance_history_file") or "balance_history").strip()
    if not name.endswith(".xlsx"):
        name += ".xlsx"
    hist_path = OUTPUT_DIR / name
    if not hist_path.exists():
        print("No balance_history file found; run the scraper once with balance capture.")
        return
    try:
        import openpyxl
    except ImportError:
        print("openpyxl required. pip install openpyxl")
        return
    latest_by_account = {}
    try:
        wb = openpyxl.load_workbook(hist_path, read_only=True, data_only=True)
        ws = wb["Balance history"] if "Balance history" in wb.sheetnames else wb.active
        headers = [ws.cell(1, c).value for c in range(1, 4)]
        snap_col = 1
        account_col = 2
        balance_col = 3
        for c, h in enumerate(headers, 1):
            if h and "snapshot" in str(h).lower():
                snap_col = c
            if h and "account" in str(h).lower():
                account_col = c
            if h and "balance" in str(h).lower() and "point" in str(h).lower():
                balance_col = c
        for row in range(2, ws.max_row + 1):
            acc = ws.cell(row, account_col).value
            bal = ws.cell(row, balance_col).value
            snap = ws.cell(row, snap_col).value
            if acc and bal is not None:
                try:
                    b = int(float(bal))
                    if row > latest_by_account.get(acc, (0, None, None))[0]:
                        latest_by_account[acc] = (row, str(snap or ""), b)
                except (TypeError, ValueError):
                    pass
        wb.close()
    except Exception as e:
        print("Could not read balance_history:", e)
        return
    for path in OUTPUT_DIR.glob("*.xlsx"):
        if path.name.startswith("~") or path.suffix.lower() != ".xlsx":
            continue
        if path.name.lower().startswith("balance_history"):
            continue
        account_name = path.stem.replace("_", " ")
        if account_name not in latest_by_account:
            continue
        _, snapshot_at, balance = latest_by_account[account_name]
        try:
            wb = openpyxl.load_workbook(path)
            if "Balance" in wb.sheetnames:
                wb.close()
                continue
            ws_bal = wb.create_sheet("Balance", 1)
            ws_bal.cell(1, 1, "Snapshot At")
            ws_bal.cell(1, 2, "Points Balance")
            ws_bal.cell(2, 1, snapshot_at)
            ws_bal.cell(2, 2, balance)
            wb.save(path)
            wb.close()
            print(f"Added Balance sheet to {path.name} (balance {balance:,})")
        except Exception as e:
            print(f"Skip {path.name}: {e}")
    print("Backfill done.")


def cleanup_existing_date_formats(config=None):
    """
    One-time cleanup: normalize all Date cells in existing account xlsx files under output/
    to YYYY-MM-DD (fixes Excel datetimes like 2025-09-15 00:00:00 and mixed formats).
    """
    config = config or load_config()
    column_order = config.get("output", {}).get("columns") or [
        "Date", "Account Name", "Payee", "Type", "EarnX", "Dollars", "Points"
    ]
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    xlsx_files = list(OUTPUT_DIR.glob("*.xlsx"))
    try:
        import openpyxl
    except ImportError:
        print("openpyxl required for date cleanup. pip install openpyxl")
        return
    for path in xlsx_files:
        if path.name.startswith("~"):
            continue
        try:
            wb = openpyxl.load_workbook(path, data_only=False)
            changed = False
            for sheet_name in wb.sheetnames:
                if sheet_name == "Balance":
                    continue
                ws = wb[sheet_name]
                if ws.max_row < 2:
                    continue
                headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
                date_col = None
                for col, h in enumerate(headers, 1):
                    if h and "date" in str(h).lower():
                        date_col = col
                        break
                if date_col is None:
                    continue
                for row in range(2, ws.max_row + 1):
                    cell = ws.cell(row=row, column=date_col)
                    val = cell.value
                    if val is None:
                        continue
                    s = str(val).strip()
                    if not s:
                        continue
                    normalized = normalize_date_str(s)
                    if normalized != s and re.match(r"^\d{4}-\d{2}-\d{2}$", normalized):
                        cell.value = normalized
                        changed = True
            if changed:
                wb.save(path)
                print(f"Cleaned dates in {path.name}")
            wb.close()
        except Exception as e:
            print(f"Skip {path.name}: {e}")
    print("Date cleanup done.")


def _normalize_row_dates(rows: list) -> None:
    """In-place: set each row's Date to canonical YYYY-MM-DD for consistent sort and dedupe."""
    for r in rows:
        d = r.get("Date")
        if d is not None and str(d).strip():
            r["Date"] = normalize_date_str(str(d).strip())

def merge_and_save(rows: list, account_name: str, column_order: list, balance: int | None = None) -> tuple:
    """
    Merge new rows into the account's ongoing file. Allows multiple rows with the same
    (Date, Payee, Dollars, Points) so legitimate duplicate transactions are all kept.
    Dates normalized to YYYY-MM-DD. If balance is provided, writes it to a 'Balance' sheet
    in the same workbook. Returns (path, num_new, num_skipped).
    """
    set_earn_x_from_points_dollars(rows)
    _normalize_row_dates(rows)
    path = _account_file_path(account_name)
    existing = load_existing_rows(path, column_order)
    set_earn_x_from_points_dollars(existing)
    _normalize_row_dates(existing)
    existing_key_counts = Counter(_row_key(r) for r in existing)
    new_rows = []
    for r in rows:
        key = _row_key(r)
        if existing_key_counts[key] > 0:
            existing_key_counts[key] -= 1
            continue
        new_rows.append(r)
    combined = existing + new_rows
    combined.sort(key=_date_sort_key)
    balance_snapshot = (datetime.now().isoformat(timespec="seconds"), balance) if balance is not None else None
    if combined:
        export_xlsx(combined, path, balance_snapshot=balance_snapshot)
        if balance_snapshot is not None:
            _ensure_balance_sheet_in_workbook(path, balance_snapshot[0], balance_snapshot[1])
    elif balance_snapshot and path.exists():
        # No transaction rows but we have a balance: update only the Balance sheet
        try:
            import openpyxl
            wb = openpyxl.load_workbook(path)
            if "Balance" in wb.sheetnames:
                ws_bal = wb["Balance"]
            else:
                ws_bal = wb.create_sheet("Balance", 1)
            ws_bal.cell(row=1, column=1, value="Snapshot At")
            ws_bal.cell(row=1, column=2, value="Points Balance")
            ws_bal.cell(row=2, column=1, value=balance_snapshot[0])
            ws_bal.cell(row=2, column=2, value=balance_snapshot[1])
            wb.save(path)
            wb.close()
        except Exception:
            pass
    return path, len(new_rows), len(rows) - len(new_rows)


def set_earn_x_from_points_dollars(rows: list) -> None:
    """
    Set EarnX from Points/Dollars when both are numeric.
    Chase sometimes shows '1x' for all; the actual multiplier is points/dollars (e.g. 3x, 2x).
    Handles Excel-loaded values that may be float/int (coerce to str).
    """
    for row in rows:
        try:
            d = str(row.get("Dollars") or "").strip().replace("$", "").replace(",", "")
            p = str(row.get("Points") or "").strip().replace(",", "")
            if not d or not p:
                continue
            dollars = float(d)
            points = float(p)
            if dollars <= 0:
                continue
            mult = points / dollars
            mult_rounded = round(mult, 1)
            if mult_rounded == int(mult_rounded):
                row["EarnX"] = f"{int(mult_rounded)}x"
            else:
                row["EarnX"] = f"{mult_rounded}x"
        except (ValueError, TypeError):
            pass


def export_csv(rows, path):
    if not rows:
        return
    keys = list(rows[0].keys())
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=keys)
        w.writeheader()
        w.writerows(rows)


def _ensure_balance_sheet_in_workbook(path: Path, snapshot_at: str, balance: int) -> None:
    """Add or update the 'Balance' sheet in an existing xlsx so the per-account file always has it when we have a balance."""
    if not path.exists() or path.suffix.lower() != ".xlsx":
        return
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path)
        if "Balance" in wb.sheetnames:
            ws_bal = wb["Balance"]
        else:
            ws_bal = wb.create_sheet("Balance", 1)
        ws_bal.cell(row=1, column=1, value="Snapshot At")
        ws_bal.cell(row=1, column=2, value="Points Balance")
        ws_bal.cell(row=2, column=1, value=snapshot_at)
        ws_bal.cell(row=2, column=2, value=balance)
        wb.save(path)
        wb.close()
    except Exception:
        pass


def export_xlsx(rows, path, balance_snapshot=None):
    """Write transaction rows to first sheet; optionally add/update 'Balance' sheet with (snapshot_at, balance)."""
    if not rows:
        return
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        export_csv(rows, path.with_suffix(".csv"))
        return
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Points"
    keys = list(rows[0].keys())
    header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for col, key in enumerate(keys, 1):
        cell = ws.cell(row=1, column=col, value=key)
        cell.fill = header_fill
        cell.font = header_font
    for row_idx, row in enumerate(rows, 2):
        for col_idx, key in enumerate(keys, 1):
            ws.cell(row=row_idx, column=col_idx, value=row.get(key, ""))
    if balance_snapshot is not None:
        snapshot_at, balance = balance_snapshot
        if "Balance" in wb.sheetnames:
            ws_bal = wb["Balance"]
        else:
            ws_bal = wb.create_sheet("Balance", 1)
        ws_bal.cell(row=1, column=1, value="Snapshot At")
        ws_bal.cell(row=1, column=2, value="Points Balance")
        ws_bal.cell(row=2, column=1, value=snapshot_at)
        ws_bal.cell(row=2, column=2, value=balance)
    wb.save(path)


def run_scraper(output_format=None, accounts=None):
    config = load_config()
    output_format = output_format or config.get("output", {}).get("default_format", "csv")
    column_order = config.get("output", {}).get("columns") or [
        "Date", "Account Name", "Payee", "Type", "EarnX", "Dollars", "Points"
    ]
    balance_path = ensure_balance_history_file(config)
    print(f"Balance history (when recorded): {balance_path}")
    sys.stdout.flush()

    if accounts is None:
        accounts = prompt_account_menu(config)
    last_rows = None

    for account_name in accounts:
        username, password = get_credentials(account_name, config)
        profile_dir, SESSION_FILE = _session_paths_for_account(account_name, config)

        headed = os.environ.get("CHASE_HEADED", "1") == "1"
        chrome = config.get("chase", {})
        login_url = chrome.get("login_url", "https://secure.chase.com/")
        login_timeout = (chrome.get("login_timeout") or 120) * 1000

        OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
        SESSION_DIR.mkdir(parents=True, exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

        profile_dir.mkdir(parents=True, exist_ok=True)
        if len(accounts) > 1:
            print("\n" + "=" * 50)
            print("Account {} of {}: {}".format(accounts.index(account_name) + 1, len(accounts), account_name))
        else:
            print("Account:", account_name)
        print("Profile (2FA saved here):", profile_dir.resolve())
        sys.stdout.flush()

        with sync_playwright() as p:
            # Persistent profile = real browser profile on disk. Chase "remember this device" cookie lives here so we only do 2FA once per account.
            context_options = {
                "viewport": {"width": 1280, "height": 900},
                "user_agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36",
                "accept_downloads": True,
                "ignore_https_errors": False,
            }
            try:
                context = p.chromium.launch_persistent_context(
                    user_data_dir=str(profile_dir),
                    channel="chrome" if headed else None,
                    headless=not headed,
                    **context_options,
                )
            except Exception:
                context = p.chromium.launch_persistent_context(
                    user_data_dir=str(profile_dir),
                    headless=not headed,
                    **context_options,
                )
            page = context.pages[0] if context.pages else context.new_page()
            page.set_default_timeout(login_timeout)

            print("Loading Chase login page...")
            sys.stdout.flush()
            try:
                page.goto(login_url, wait_until="domcontentloaded", timeout=30000)
                if "system-requirements" in page.url:
                    print("Redirecting from system-requirements page...")
                    page.goto(login_url, wait_until="domcontentloaded")
                page.wait_for_load_state("networkidle", timeout=15000)
            except PlaywrightTimeout:
                pass
            print("Waiting for login form...")
            sys.stdout.flush()
            page.wait_for_timeout(3000)

            # Check if we need to log in (login form can be in main page or in an iframe)
            need_login = False
            try:
                content = (page.content() or "").lower()
                if "confirm your identity" in content:
                    need_login = True
                if not need_login:
                    # Main document
                    if page.query_selector("input[name='userId'], input[id*='userId'], input[type='password']"):
                        need_login = True
                if not need_login:
                    # Login form is often inside an iframe — check every frame
                    for frame in page.frames:
                        try:
                            if frame.locator("input[type='password']").first.is_visible(timeout=500):
                                need_login = True
                                break
                            if frame.locator("input[name='userId'], input[id*='userId'], input[type='text']").first.is_visible(timeout=300):
                                need_login = True
                                break
                        except Exception:
                            continue
                # If we're on Chase and URL doesn't look like post-login, assume login needed
                if not need_login and "chase.com" in page.url and "dashboard" not in page.url and "overview" not in page.url and "account" not in page.url:
                    if "sign in" in content or "username" in content or "password" in content:
                        need_login = True
            except Exception:
                need_login = True

            if need_login:
                # Save login page for debugging (so we can see exact DOM/iframes)
                try:
                    debug_path = OUTPUT_DIR / "login_page_debug.html"
                    debug_path.write_text(page.content(), encoding="utf-8")
                    print("Saved login page HTML to", debug_path)
                    sys.stdout.flush()
                except Exception:
                    pass
    
                # Use Playwright frame API: find the frame with the login form, then fill username then password (no JS — works with any iframe)
                print("Looking for login form (username + password)...")
                sys.stdout.flush()
                filled = False
                for frame in page.frames:
                    try:
                        pwd = frame.locator("input[type='password']").first
                        if not pwd.is_visible(timeout=1500):
                            continue
                        # In this frame: first non-password input = username (then password)
                        user_input = frame.locator("input:not([type='password']):not([type='hidden'])").first
                        user_input.wait_for(state="visible", timeout=5000)
                        user_input.click()
                        page.wait_for_timeout(100)
                        user_input.fill(username)
                        print("Filled username.")
                        sys.stdout.flush()
                        page.wait_for_timeout(200)
                        pwd.click()
                        page.wait_for_timeout(100)
                        pwd.fill(password)
                        print("Filled password.")
                        sys.stdout.flush()
                        page.wait_for_timeout(300)
                        btn = frame.locator("button:has-text('Sign in'), input[type='submit'], button[type='submit']").first
                        btn.click()
                        print("Clicked Sign in.")
                        sys.stdout.flush()
                        filled = True
                        break
                    except Exception:
                        continue
                if not filled:
                    print("Could not find login form in any frame; enter credentials manually.")
                    sys.stdout.flush()
                page.wait_for_timeout(5000)
    
                # Only skip 2FA wait when we see the real main page (dashboard/account selector), not just the URL
                main_page_loaded = False
                for _ in range(25):
                    page.wait_for_timeout(1000)
                    content = (page.content() or "").lower()
                    url = (page.url or "").lower()
                    has_2fa_prompt = (
                        "confirm your identity" in content
                        or "verify your identity" in content
                        or "enter the code" in content
                        or "one-time code" in content
                    )
                    if has_2fa_prompt:
                        main_page_loaded = False
                        break
                    # Real dashboard/overview has these; 2FA page does not (no time-sensitive greeting)
                    on_dashboard = (
                        "bank accounts" in content
                        or "credit cards" in content
                        or ("sign out" in content and ("overview" in content or "accounts" in content))
                    )
                    on_account_selector = "account-selector" in url and "choose" in content
                    if on_dashboard or on_account_selector:
                        main_page_loaded = True
                        print("Main page loaded - no 2FA required.")
                        sys.stdout.flush()
                        break
                if not main_page_loaded:
                    _prompt_browser(
                        "Complete 'Confirm Your Identity' — choose Get a text or Get a call and enter the code.",
                        "When you see the Chase dashboard (or account selector), press ENTER in this terminal to continue.",
                    )
                    sys.stdout.flush()
                    input(">>> Press ENTER after completing 2FA... ")
                    print("Continuing...")
                    sys.stdout.flush()
    
                try:
                    page.wait_for_url(re.compile(r"chase\.com", re.I), timeout=login_timeout)
                    page.wait_for_load_state("networkidle", timeout=30000)
                    # Wait until we're past the identity page (e.g. dashboard) so session is fully established
                    for _ in range(30):
                        page.wait_for_timeout(1000)
                        content = (page.content() or "").lower()
                        if "confirm your identity" not in content and ("dashboard" in page.url or "overview" in page.url or "account" in page.url):
                            break
                except PlaywrightTimeout:
                    print("Continuing after timeout; ensure you're on the rewards/activity page.")
    
                try:
                    SESSION_FILE.parent.mkdir(parents=True, exist_ok=True)
                    context.storage_state(path=str(SESSION_FILE))
                    print("Session saved (2FA). Next run may skip login:", SESSION_FILE)
                except Exception as e:
                    print("Could not save session:", e)
                    sys.stdout.flush()
            else:
                print("Using saved session - already logged in.")
                sys.stdout.flush()
    
            # Wait for dashboard if needed, then go to account-selector and select this account by last 4 digits
            dashboard_url = chrome.get("dashboard_url") or "https://secure.chase.com/web/auth/dashboard#/dashboard/overview"
            selector_url = chrome.get("account_selector_url") or "https://ultimaterewardspoints.chase.com/account-selector"
            last4 = (config.get("account_last4") or {}).get(account_name, "").strip()
            balance_from_selector = None  # captured on selector page before we navigate to activity

            try:
                if "dashboard" not in page.url and "overview" not in page.url:
                    print("Waiting for dashboard...")
                    sys.stdout.flush()
                    page.goto(dashboard_url, wait_until="load", timeout=30000)
                    try:
                        page.wait_for_load_state("networkidle", timeout=8000)
                    except Exception:
                        pass
                print("Going to account selector...")
                sys.stdout.flush()
                page.goto(selector_url, wait_until="load", timeout=30000)
                try:
                    page.wait_for_load_state("networkidle", timeout=3000)
                except Exception:
                    pass
                page.wait_for_timeout(1000)
                try:
                    page.wait_for_url(re.compile(r"account-selector|ultimaterewardspoints"), timeout=3000)
                except Exception:
                    pass
    
                # If Chase is asking for 2FA (e.g. session was stale), pause for user to complete it
                if "confirm your identity" in (page.content() or "").lower():
                    _prompt_browser(
                        "Chase is asking to confirm your identity. Complete 2FA in the browser.",
                        "When you see the account selector or dashboard, press ENTER here to continue.",
                    )
                    sys.stdout.flush()
                    input(">>> Press ENTER after completing 2FA... ")
                    print("Continuing...")
                    sys.stdout.flush()
                    page.wait_for_timeout(1000)
    
                if last4:
                    # Exact structure from Chase: body.account-selector-container > section.card-section > mds-list > mds-list-item[label="...1827"][href="..."]
                    selected = False
                    try:
                        # "visible" often times out on mds-list-item; "attached" is enough (balance is in description attr)
                        page.wait_for_selector(
                            "section.card-section mds-list-item, body.account-selector-container mds-list-item",
                            state="attached",
                            timeout=15000,
                        )
                        page.wait_for_timeout(500)
                    except Exception as ex:
                        print(f"[Balance] Selector wait (attached): {ex}")
                        sys.stdout.flush()
                    print(f"[Balance] Reading from selector page (last4={last4})...")
                    sys.stdout.flush()
                    try:
                        balance_from_selector = scrape_balance_from_selector_page(page, last4)
                    except Exception as ex:
                        print(f"[Balance] Selector scrape error: {ex}")
                        sys.stdout.flush()
                        balance_from_selector = None
                    if balance_from_selector is not None:
                        print(f"[Balance] Selector page returned: {balance_from_selector:,}")
                    else:
                        print("[Balance] Selector page returned: None")
                    sys.stdout.flush()
                    # Primary: mds-list-item has label and href in light DOM — get href and navigate (no click needed)
                    try:
                        card = page.locator(f'mds-list-item[label*="{last4}"]').first
                        card.wait_for(state="attached", timeout=10000)
                        if balance_from_selector is None:
                            balance_from_selector = scrape_balance_from_selector_page(page, last4)
                        href = card.get_attribute("href")
                        if href:
                            page.goto(href, wait_until="load", timeout=20000)
                            try:
                                page.wait_for_load_state("networkidle", timeout=3000)
                            except Exception:
                                pass
                            print("Selected account ending in", last4)
                            sys.stdout.flush()
                            selected = True
                    except Exception as e1:
                        pass
                    if not selected:
                        try:
                            href_or_clicked = page.evaluate("""(last4) => {
                                const items = document.querySelectorAll('section.card-section mds-list-item');
                                for (const el of items) {
                                    const label = el.getAttribute('label') || '';
                                    if (label.indexOf(last4) !== -1) {
                                        const href = el.getAttribute('href') || '';
                                        if (href) return { href: href };
                                        if (el.shadowRoot) {
                                            const a = el.shadowRoot.querySelector('a[href*="ultimaterewardspoints"]');
                                            if (a) { a.click(); return { clicked: true }; }
                                        }
                                        el.click();
                                        return { clicked: true };
                                    }
                                }
                                return null;
                            }""", last4)
                            if href_or_clicked and href_or_clicked.get("href"):
                                page.goto(href_or_clicked["href"], wait_until="load", timeout=20000)
                                try:
                                    page.wait_for_load_state("networkidle", timeout=3000)
                                except Exception:
                                    pass
                                print("Selected account ending in", last4, "(JS)")
                                sys.stdout.flush()
                                selected = True
                            elif href_or_clicked and href_or_clicked.get("clicked"):
                                page.wait_for_timeout(2000)
                                selected = True
                        except Exception as e0:
                            print("JS selection:", e0)
                            sys.stdout.flush()
                    if not selected:
                        try:
                            link = page.get_by_role("link", name=re.compile(re.escape(last4)))
                            link.first.wait_for(state="visible", timeout=6000)
                            link.first.click()
                            print("Selected account ending in", last4, "(link)")
                            sys.stdout.flush()
                            page.wait_for_timeout(2000)
                            selected = True
                        except Exception as e2:
                            print("Could not select account for digits", last4, "-", e2)
                            sys.stdout.flush()
                            print(">>> Please click the account card in the browser, then press ENTER here to continue.")
                            sys.stdout.flush()
                            input()
                else:
                    print("No account_last4 for", account_name, "in config - select account manually.")
                    sys.stdout.flush()
            except Exception as e:
                print("Navigation to account selector failed:", e)
                sys.stdout.flush()
    
            # Auto-navigate: click "See rewards activity" then "See all transactions" (keep Enter prompt as fallback)
            rewards_url = chrome.get("rewards_activity_url")
            if rewards_url:
                print("Navigating to rewards activity URL...")
                sys.stdout.flush()
                page.goto(rewards_url, wait_until="load", timeout=30000)
            else:
                try:
                    print("Clicking 'See rewards activity'...")
                    sys.stdout.flush()
                    btn = (
                        page.get_by_role("button", name=re.compile(r"See rewards activity", re.I))
                        .or_(page.get_by_role("link", name=re.compile(r"See rewards activity", re.I)))
                        .or_(page.get_by_text("See rewards activity", exact=False))
                        .first
                    )
                    btn.wait_for(state="visible", timeout=10000)
                    btn.click()
                    page.wait_for_load_state("load", timeout=15000)
                    page.wait_for_timeout(2000)
                except Exception as e1:
                    print("Could not click See rewards activity:", e1)
                    sys.stdout.flush()
                try:
                    print("Clicking 'See all transactions'...")
                    sys.stdout.flush()
                    btn2 = (
                        page.get_by_role("button", name=re.compile(r"See all transactions", re.I))
                        .or_(page.get_by_role("link", name=re.compile(r"See all transactions", re.I)))
                        .or_(page.get_by_text("See all transactions", exact=False))
                        .first
                    )
                    btn2.wait_for(state="visible", timeout=10000)
                    btn2.click()
                    page.wait_for_load_state("load", timeout=15000)
                    page.wait_for_timeout(2000)
                except Exception as e2:
                    print("Could not click See all transactions:", e2)
                    sys.stdout.flush()
    
            # Click "See more transactions" until it becomes "Back to top" (all transactions loaded)
            max_load_more = 200
            for n in range(max_load_more):
                see_more = (
                    page.get_by_role("button", name=re.compile(r"See more transactions", re.I))
                    .or_(page.get_by_role("link", name=re.compile(r"See more transactions", re.I)))
                    .or_(page.get_by_text("See more transactions", exact=False))
                    .first
                )
                try:
                    if see_more.is_visible(timeout=2000):
                        see_more.click()
                        page.wait_for_timeout(1500)
                        if (n + 1) % 10 == 0 or n < 3:
                            print("Loaded more transactions...", n + 1)
                            sys.stdout.flush()
                        continue
                except Exception:
                    pass
                back_to_top = page.get_by_role("button", name=re.compile(r"Back to top", re.I)).or_(page.get_by_text("Back to top", exact=False)).first
                try:
                    if back_to_top.is_visible(timeout=1000):
                        print("All transactions loaded (Back to top visible).")
                        sys.stdout.flush()
                        break
                except Exception:
                    pass
                break
            else:
                print("Reached max 'See more' clicks; continuing with current list.")
                sys.stdout.flush()
    
            print("Scraping current page for points activity table...")
            sys.stdout.flush()
            page.wait_for_timeout(5000)  # allow dynamic content (e.g. React) to render
            rows, _ = scrape_table_and_account(page, config)
            if rows:
                rows = add_account_column(rows, account_name)
            rows = ensure_column_order(rows, column_order)

            balance = scrape_points_balance(page, config)
            if balance is not None:
                print(f"[Balance] From activity page config/fallback: {balance:,}")
                sys.stdout.flush()
            if balance is None and balance_from_selector is not None:
                balance = balance_from_selector
                print(f"[Balance] From selector page (saved earlier): {balance:,}")
                sys.stdout.flush()
            if balance is None:
                print("[Balance] Trying activity page shadow-DOM scan...")
                sys.stdout.flush()
                balance = scrape_balance_from_activity_page(page)
                if balance is not None:
                    print(f"[Balance] From activity page: {balance:,}")
                    sys.stdout.flush()
            if balance is not None:
                bp = append_balance_snapshot(account_name, balance, config)
                if bp:
                    print(f"Points balance {balance:,} recorded -> {bp}")
                    sys.stdout.flush()
            elif rows:
                print(
                    "Could not read points balance. Add selectors.points_balance in config.yaml "
                    "(CSS to the element showing total points) if this persists."
                )
                sys.stdout.flush()

            if not rows:
                debug_path = OUTPUT_DIR / "last_page_debug.html"
                try:
                    debug_path.write_text(page.content(), encoding="utf-8")
                    print(f"No activity table found. Page HTML saved to {debug_path} for inspection.")
                except Exception:
                    print("No activity table found. Navigate to your Chase rewards/points activity page and run again, or check config.yaml selectors.")
                print("Tip: Next run, open the rewards/points activity page in the browser within 90 seconds so the script can detect it.")
            else:
                print(f"Scraped {len(rows)} rows. Account: {account_name}")
    
            # Backup cookies to JSON (profile is the source of truth for 2FA)
            try:
                SESSION_FILE.parent.mkdir(parents=True, exist_ok=True)
                context.storage_state(path=str(SESSION_FILE))
            except Exception:
                pass
    
            # Sign out of Chase before closing (clear session for next run)
            try:
                print("Signing out...")
                sys.stdout.flush()
                try:
                    page.goto("https://secure.chase.com/web/auth/logout", wait_until="domcontentloaded", timeout=10000)
                    page.wait_for_timeout(2000)
                except Exception:
                    pass
                try:
                    sign_out = (
                        page.get_by_role("link", name=re.compile(r"Sign out|Log out|Log off", re.I))
                        .or_(page.get_by_role("button", name=re.compile(r"Sign out|Log out|Log off", re.I)))
                        .or_(page.get_by_text(re.compile(r"Sign out|Log out|Log off", re.I)))
                        .first
                    )
                    if sign_out.is_visible(timeout=2000):
                        sign_out.click()
                        page.wait_for_timeout(2000)
                except Exception:
                    pass
                print("Signed out.")
                sys.stdout.flush()
            except Exception as e:
                print("Sign out skipped:", e)
                sys.stdout.flush()
    
            context.close()

        # Merge into ongoing file for this account (one file per account); include balance in workbook if we have it
        if rows:
            last_rows = rows
            out_path, num_new, num_skipped = merge_and_save(rows, account_name, column_order, balance=balance)
            print(f"Saved to {out_path}: {num_new} new transaction(s) added, {num_skipped} already present.")
            sys.stdout.flush()

    return last_rows


if __name__ == "__main__":
    if len(sys.argv) > 1:
        arg = sys.argv[1].strip().lower()
        if arg in ("--cleanup-dates", "-c"):
            print("One-time date cleanup: normalizing Date column to YYYY-MM-DD in all account xlsx files...")
            cleanup_existing_date_formats()
        elif arg in ("--backfill-balance", "-b"):
            print("Backfilling Balance sheet into account workbooks from balance_history.xlsx...")
            backfill_balance_sheets()
        else:
            run_scraper()
    else:
        run_scraper()
